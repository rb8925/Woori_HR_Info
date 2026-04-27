import os

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from data_parser import (
    COMPANIES,
    PREV_YEAR,
    RECENT_YEAR,
    TREND_YEARS,
    build_table1,
    build_table2,
    fetch_all_raw,
)

# ── 페이지 설정 ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="증권사 인력 현황 비교",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

HIGHLIGHT_CO   = "우리투자증권"
HIGHLIGHT_BG   = "#FFF8DC"   # 연한 노란색
HEADER_BG      = "#1A5276"

GROUPS_DEF: dict[str, list[str]] = {
    "대형사":    ["미래에셋증권", "한국투자증권", "NH투자증권", "KB증권"],
    "중·대형사": ["키움증권", "메리츠증권"],
    "중·소형사": ["유안타증권", "현대차증권", "IBK투자증권", "유진투자증권"],
}
AVG_COL_NAME: dict[str, str] = {g: f"{g} 평균" for g in GROUPS_DEF}
AVG_COLS: set[str] = set(AVG_COL_NAME.values())

COMPANY_GROUP: dict[str, str] = {
    "우리투자증권":              "",
    "미래에셋증권":              "대형사",
    "한국투자증권":              "대형사",
    "NH투자증권":                "대형사",
    "KB증권":                    "대형사",
    AVG_COL_NAME["대형사"]:      "대형사",
    "키움증권":                  "중·대형사",
    "메리츠증권":                "중·대형사",
    AVG_COL_NAME["중·대형사"]:   "중·대형사",
    "유안타증권":                "중·소형사",
    "현대차증권":                "중·소형사",
    "IBK투자증권":               "중·소형사",
    "유진투자증권":              "중·소형사",
    AVG_COL_NAME["중·소형사"]:   "중·소형사",
}
GROUP_BG = "#EBF5FB"

# ── 숫자 포맷 ─────────────────────────────────────────────────────────────────
def _fmt(v, kind: str) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "-"
    try:
        if kind == "int":
            return f"{int(round(v)):,}"
        if kind == "pct":
            return f"{v:,.1f}%"
        # decimal
        return f"{v:,.1f}"
    except Exception:
        return "-"

# 행(지표)별 포맷 타입 정의
_ROW_FMT: dict[str, str] = {
    "자기자본(억원)":       "int",
    "리테일 인원":          "int",
    "리테일 비중(%)":       "pct",
    "본사영업 인원":        "int",
    "본사영업 비중(%)":     "pct",
    "본사관리 인원":        "int",
    "본사관리 비중(%)":     "pct",
    "총인원":               "int",
    "인당생산성(백만원/인)": "decimal",
    "임직원수 평균":         "int",
    "순영업수익 평균(억원)": "int",
    **{f"임직원수 {y}년":        "int" for y in TREND_YEARS},
    **{f"순영업수익 {y}년(억원)": "int" for y in TREND_YEARS},
}

def format_df(df: pd.DataFrame) -> pd.DataFrame:
    """DataFrame의 모든 값을 표시용 문자열로 변환."""
    out = df.copy().astype(object)
    for row_label in out.index:
        kind = _ROW_FMT.get(row_label, "decimal")
        out.loc[row_label] = out.loc[row_label].apply(lambda v: _fmt(v, kind))
    return out


def _merge_pct_rows(df: pd.DataFrame) -> pd.DataFrame:
    """인원 행과 비중 행을 '1,213 (36.8%)' 형식으로 합쳐 단일 행으로 만듦 (포맷된 문자열 DataFrame에 적용)."""
    pairs = [
        ("리테일 인원",   "리테일 비중(%)"),
        ("본사영업 인원", "본사영업 비중(%)"),
        ("본사관리 인원", "본사관리 비중(%)"),
    ]
    skip = {p for _, p in pairs}
    pair_map = {h: p for h, p in pairs}

    rows_data: dict[str, dict] = {}
    for lbl in df.index:
        if lbl in skip:
            continue
        pct_lbl = pair_map.get(lbl)
        if pct_lbl and pct_lbl in df.index:
            merged: dict[str, str] = {}
            for col in df.columns:
                hv, pv = df.loc[lbl, col], df.loc[pct_lbl, col]
                if hv == "-":
                    merged[col] = "-"
                elif pv == "-":
                    merged[col] = hv
                else:
                    merged[col] = f"{hv} ({pv})"
            rows_data[lbl] = merged
        else:
            rows_data[lbl] = df.loc[lbl].to_dict()

    return pd.DataFrame(rows_data).T


def add_group_averages(df: pd.DataFrame) -> pd.DataFrame:
    """각 그룹 마지막 열 우측에 그룹 평균 열 삽입 (수치형 DataFrame)."""
    out = df.copy()
    cols_ordered: list[str] = []
    for co in df.columns:
        cols_ordered.append(co)
        for g, members in GROUPS_DEF.items():
            if co == members[-1]:
                avg_col = AVG_COL_NAME[g]
                valid = [c for c in members if c in df.columns]
                out[avg_col] = df[valid].mean(axis=1, skipna=True)
                cols_ordered.append(avg_col)
    return out[cols_ordered]

def _render_table(df_fmt: pd.DataFrame, section_rows: bool = False) -> str:
    """colspan/rowspan이 적용된 HTML 테이블 생성."""
    companies = list(df_fmt.columns)

    # 연속된 같은 그룹으로 컬럼 묶기
    col_groups: list[list] = []
    for co in companies:
        g = COMPANY_GROUP.get(co, "")
        if col_groups and col_groups[-1][0] == g:
            col_groups[-1][1].append(co)
        else:
            col_groups.append([g, [co]])

    PROD_BG = "#E9F7EF"   # 인당생산성 행 배경색
    AVG_BG  = "#EEF2F7"   # 그룹 평균 열 배경색
    S = {
        "tbl":        "border-collapse:collapse;width:100%;font-size:13px;font-family:sans-serif;",
        "th":         "border:1px solid #ccc;padding:6px 10px;font-weight:600;white-space:nowrap;text-align:center;background:#F2F3F4;",
        "th_hl":      "border:1px solid #ccc;padding:6px 10px;font-weight:600;white-space:nowrap;text-align:center;background:#FFF8DC;",
        "th_grp":     "border:1px solid #ccc;padding:6px 10px;font-weight:700;white-space:nowrap;text-align:center;background:#D6EAF8;",
        "th_idx":     "border:1px solid #ccc;padding:6px 10px;font-weight:700;white-space:nowrap;text-align:center;background:#F2F3F4;",
        "th_avg":     f"border:1px solid #ccc;padding:6px 10px;font-weight:600;white-space:nowrap;text-align:center;background:{AVG_BG};font-style:italic;",
        "td":         "border:1px solid #ccc;padding:5px 10px;text-align:right;white-space:nowrap;",
        "td_hl":      "border:1px solid #ccc;padding:5px 10px;text-align:right;white-space:nowrap;background:#FFF8DC;font-weight:600;",
        "td_avg":     f"border:1px solid #ccc;padding:5px 10px;text-align:right;white-space:nowrap;background:{AVG_BG};font-style:italic;",
        "td_row":     "border:1px solid #ccc;padding:5px 10px;font-weight:600;text-align:left;white-space:nowrap;",
        "td_sec":     "border:1px solid #ccc;padding:5px 10px;font-weight:700;text-align:center;vertical-align:middle;background:#EBF5FB;white-space:nowrap;",
        "td_prod":    f"border:1px solid #ccc;padding:5px 10px;font-weight:700;text-align:left;white-space:nowrap;background:{PROD_BG};",
        "td_prod_d":  f"border:1px solid #ccc;padding:5px 10px;text-align:right;white-space:nowrap;background:{PROD_BG};",
        "td_prod_hl": f"border:1px solid #ccc;padding:5px 10px;text-align:right;white-space:nowrap;background:{PROD_BG};font-weight:600;",
        "td_prod_avg":f"border:1px solid #ccc;padding:5px 10px;text-align:right;white-space:nowrap;background:{PROD_BG};font-style:italic;",
    }

    n_idx = 2 if section_rows else 1

    # ── thead ──────────────────────────────────────────────────────────────
    r1, r2 = [], []
    r1.append(f'<th rowspan="2" colspan="{n_idx}" style="{S["th_idx"]}">구분</th>')
    for g, cos in col_groups:
        if not g:   # 그룹 없음 → 2행 병합
            for co in cos:
                s = S["th_hl"] if co == HIGHLIGHT_CO else S["th"]
                r1.append(f'<th rowspan="2" style="{s}">{co}</th>')
        else:       # 그룹 있음 → 상단에 그룹명(colspan), 하단에 회사명
            r1.append(f'<th colspan="{len(cos)}" style="{S["th_grp"]}">{g}</th>')
            for co in cos:
                if co in AVG_COLS:
                    s = S["th_avg"]
                elif co == HIGHLIGHT_CO:
                    s = S["th_hl"]
                else:
                    s = S["th"]
                display = "평균" if co in AVG_COLS else co
                r2.append(f'<th style="{s}">{display}</th>')
    thead = f'<thead><tr>{"".join(r1)}</tr><tr>{"".join(r2)}</tr></thead>'

    # ── tbody ──────────────────────────────────────────────────────────────
    def _td(val, co, prod=False):
        is_avg = co in AVG_COLS
        if prod:
            s = S["td_prod_avg"] if is_avg else S["td_prod_hl"] if co == HIGHLIGHT_CO else S["td_prod_d"]
        else:
            s = S["td_avg"] if is_avg else S["td_hl"] if co == HIGHLIGHT_CO else S["td"]
        return f'<td style="{s}">{val}</td>'

    body_rows = []
    labels = list(df_fmt.index)

    if section_rows:
        def _sec(lbl):
            if "임직원수" in lbl:   return "임직원수"
            if "순영업수익" in lbl: return "순영업수익"
            return ""

        sec_groups: list[list] = []
        for lbl in labels:
            s = _sec(lbl)
            if sec_groups and sec_groups[-1][0] == s:
                sec_groups[-1][1].append(lbl)
            else:
                sec_groups.append([s, [lbl]])

        import re as _re
        SEC_DISPLAY = {"임직원수": "임직원수 (명)", "순영업수익": "순영업수익 (억원)"}

        def _disp(lbl: str, sec: str) -> str:
            """임직원수/순영업수익 섹션의 행 레이블을 연월 형식으로 단순화."""
            if sec in ("임직원수", "순영업수익"):
                if "평균" in lbl:
                    return "평균"
                m = _re.search(r"(\d{4})", lbl)
                return f"{m.group(1)[2:]}년 12월" if m else lbl
            return lbl

        for sec, lbls in sec_groups:
            is_prod = (sec == "")
            for j, lbl in enumerate(lbls):
                cells = []
                if is_prod:
                    # 인당생산성: 좌측 두 셀을 하나로 병합
                    cells.append(f'<td colspan="2" style="{S["td_prod"]}">{_disp(lbl, sec)}</td>')
                else:
                    if j == 0:
                        sec_hdr = SEC_DISPLAY.get(sec, sec)
                        cells.append(f'<td rowspan="{len(lbls)}" style="{S["td_sec"]}">{sec_hdr}</td>')
                    cells.append(f'<td style="{S["td_row"]}">{_disp(lbl, sec)}</td>')
                cells += [_td(df_fmt.loc[lbl, co], co, prod=is_prod) for co in companies]
                body_rows.append(f'<tr>{"".join(cells)}</tr>')
    else:
        for lbl in labels:
            cells = [f'<td style="{S["td_row"]}">{lbl}</td>']
            cells += [_td(df_fmt.loc[lbl, co], co) for co in companies]
            body_rows.append(f'<tr>{"".join(cells)}</tr>')

    tbody = f'<tbody>{"".join(body_rows)}</tbody>'
    return f'<div style="overflow-x:auto;"><table style="{S["tbl"]}">{thead}{tbody}</table></div>'


def _mark_partial(df: pd.DataFrame, fin_data: dict) -> pd.DataFrame:
    """판관비 누락으로 부분 계산된 순영업수익 셀에 '*' 표시."""
    df = df.copy()

    # 회사별 판관비 누락 연도 수집
    missing_by_co: dict[str, list[int]] = {}
    for company in COMPANIES:
        missing = [
            y for y in TREND_YEARS
            if fin_data.get(company, {}).get(y) is not None
            and fin_data.get(company, {}).get(y, {}).get("판관비") is None
        ]
        if missing:
            missing_by_co[company] = missing

    def _apply(col: str, missing_years: list[int]):
        if col not in df.columns:
            return
        for y in missing_years:
            lbl = f"순영업수익  {y}년 (억원)"
            if lbl in df.index and df.loc[lbl, col] != "-":
                df.loc[lbl, col] += "*"
        avg_lbl = "순영업수익  평균 (억원)"
        if avg_lbl in df.index and df.loc[avg_lbl, col] != "-":
            df.loc[avg_lbl, col] += "*"
        if RECENT_YEAR in missing_years:
            prod_lbl = "인당생산성 (백만원/인)"
            if prod_lbl in df.index and df.loc[prod_lbl, col] != "-":
                df.loc[prod_lbl, col] += "*"

    # 개별 회사 마킹
    for co, missing in missing_by_co.items():
        _apply(co, missing)

    # 그룹 평균 열: 멤버 중 하나라도 누락 있으면 동일하게 마킹
    for g, members in GROUPS_DEF.items():
        all_missing = sorted({y for co in members for y in missing_by_co.get(co, [])})
        if all_missing:
            _apply(AVG_COL_NAME[g], all_missing)

    return df


def _apply_woori_override(t1_recent: pd.DataFrame, t1_prev: pd.DataFrame,
                          t2: pd.DataFrame, ov: dict) -> tuple:
    """우리투자증권 수동 입력값을 DataFrame에 반영하고 평균·생산성을 재계산."""
    co = "우리투자증권"

    for raw, yr in [(t1_recent, RECENT_YEAR), (t1_prev, PREV_YEAR)]:
        d = ov.get(yr, {})
        if d.get("자기자본") is not None:
            raw.loc["자기자본(억원)", co] = float(d["자기자본"])
        if d.get("총인원") is not None:
            raw.loc["총인원", co] = int(d["총인원"])

    for yr in TREND_YEARS:
        d = ov.get(yr, {})
        if d.get("총인원") is not None:
            t2.loc[f"임직원수 {yr}년", co] = int(d["총인원"])
        if d.get("순영업수익") is not None:
            t2.loc[f"순영업수익 {yr}년(억원)", co] = float(d["순영업수익"])

    def _avg2(vals):
        valid = [v for v in vals
                 if v is not None and not (isinstance(v, float) and pd.isna(v)) and v != 0]
        return round(sum(valid) / len(valid), 0) if valid else None

    emp_vals = [t2.loc[f"임직원수 {y}년", co] for y in TREND_YEARS]
    nor_vals = [t2.loc[f"순영업수익 {y}년(억원)", co] for y in TREND_YEARS]
    emp_avg  = _avg2(emp_vals)
    nor_avg  = _avg2(nor_vals)
    t2.loc["임직원수 평균", co]          = emp_avg
    t2.loc["순영업수익 평균(억원)", co]  = nor_avg
    if emp_avg and nor_avg:
        # nor_avg(억원) × 1e8 / emp_avg / 1e6 = nor_avg × 100 / emp_avg
        t2.loc["인당생산성(백만원/인)", co] = round(float(nor_avg) * 100 / float(emp_avg), 1)

    return t1_recent, t1_prev, t2


# ── 데이터 로드 (캐싱) ────────────────────────────────────────────────────────
@st.cache_data(show_spinner="DART에서 데이터를 불러오는 중...")
def _load(force_refresh: bool = False):
    return fetch_all_raw(use_cache=not force_refresh)

# ── 사이드바 ──────────────────────────────────────────────────────────────────
def _refresh_password() -> str:
    try:
        import streamlit as _st
        return _st.secrets.get("REFRESH_PASSWORD", "")
    except Exception:
        return os.getenv("REFRESH_PASSWORD", "")

with st.sidebar:
    st.title("설정")

    with st.expander("데이터 재수집 (관리자)"):
        pw_input = st.text_input("비밀번호", type="password", key="refresh_pw")
        correct_pw = _refresh_password()
        if correct_pw and pw_input == correct_pw:
            if st.button("DART 데이터 재수집", type="primary", use_container_width=True):
                _load.clear()
                if os.path.exists("data/raw_cache.json"):
                    os.remove("data/raw_cache.json")
                st.rerun()
        elif pw_input:
            st.error("비밀번호가 틀렸습니다.")

    st.divider()
    st.markdown(f"""
**기준연도**: {RECENT_YEAR}년
**추이 기간**: {TREND_YEARS[-1]}~{TREND_YEARS[0]}년
**재무제표**: 개별(OFS)
**출처**: DART OpenAPI
""")
    st.divider()
    st.markdown("""
**데이터 유의사항**
- 우리투자증권: 2024년이 첫 사업보고서 (이전 연도 N/A)
- 메리츠증권: 판관비 미공시 → 순영업수익 N/A
- 순영업수익 2022년: DART API 미제공 (전사 N/A)
- 우리투자증권 섹션 구분: DART 미신고 (총인원만 표시)
""")

    st.divider()
    woori_ov: dict[int, dict] = {}
    with st.expander("우리투자증권 직접입력"):
        st.caption("0 입력 시 DART 데이터 그대로 사용")
        for y in TREND_YEARS:
            st.markdown(f"**{y}년**")
            d: dict = {}
            if y in (RECENT_YEAR, PREV_YEAR):
                v = st.number_input("자기자본 (억원)", min_value=0, value=0,
                                    step=100, key=f"ov_eq_{y}")
                d["자기자본"] = int(v) if v else None
            v = st.number_input("총인원 (명)", min_value=0, value=0,
                                step=10, key=f"ov_hc_{y}")
            d["총인원"] = int(v) if v else None
            v = st.number_input("순영업수익 (억원)", min_value=0, value=0,
                                step=10, key=f"ov_nr_{y}")
            d["순영업수익"] = float(v) if v else None
            woori_ov[y] = d

# ── 타이틀 ────────────────────────────────────────────────────────────────────
st.title("증권사 인력 현황 비교 대시보드")
st.caption(f"기준: {RECENT_YEAR}년 사업보고서 | 개별재무제표 | DART OpenAPI  ·  "
           f"노란색 열 = {HIGHLIGHT_CO}")

# ── 데이터 로드 ───────────────────────────────────────────────────────────────
emp_data, fin_data = _load()
t1_raw      = build_table1(emp_data, fin_data, RECENT_YEAR)
t1_raw_prev = build_table1(emp_data, fin_data, PREV_YEAR)
t2_raw      = build_table2(emp_data, fin_data)
t1_raw, t1_raw_prev, t2_raw = _apply_woori_override(t1_raw, t1_raw_prev, t2_raw, woori_ov)

T1_ROW_LABELS = {
    "자기자본(억원)":   "자기자본 (억원)",
    "리테일 인원":      "리테일 인원",
    "본사영업 인원":    "본사영업 인원",
    "본사관리 인원":    "본사관리 인원",
    "총인원":           "총인원",
}

def _build_t1_html(raw, year):
    fmt = (
        _merge_pct_rows(format_df(add_group_averages(raw)))
        .rename(index=T1_ROW_LABELS)
    )
    return _render_table(fmt)

# ── Table 1 (2025) ────────────────────────────────────────────────────────────
st.subheader(f"증권사별 자기자본 및 인력 현황 ({RECENT_YEAR}년)")
st.html(_build_t1_html(t1_raw, RECENT_YEAR))

st.divider()

# ── Table 1 (2024) ────────────────────────────────────────────────────────────
st.subheader(f"증권사별 자기자본 및 인력 현황 ({PREV_YEAR}년)")
st.html(_build_t1_html(t1_raw_prev, PREV_YEAR))

st.divider()

# ── 인력 비교 꺾은선 그래프 ──────────────────────────────────────────────────
st.subheader(f"{PREV_YEAR} vs {RECENT_YEAR} 인력 현황 비교")

METRIC_OPTIONS = {
    "총원":     "총인원",
    "리테일":   "리테일 인원",
    "본사영업": "본사영업 인원",
    "본사관리": "본사관리 인원",
}

selected_metric = st.radio(
    "기준",
    list(METRIC_OPTIONS.keys()),
    horizontal=True,
    label_visibility="collapsed",
)
row_key = METRIC_OPTIONS[selected_metric]
cos     = list(COMPANIES.keys())

def _get_vals(raw: pd.DataFrame, key: str) -> list:
    if key not in raw.index:
        return [None] * len(cos)
    return [
        (float(raw.loc[key, co]) if raw.loc[key, co] is not None and not (isinstance(raw.loc[key, co], float) and pd.isna(raw.loc[key, co])) else None)
        for co in cos
    ]

vals_prev = _get_vals(t1_raw_prev, row_key)
vals_cur  = _get_vals(t1_raw, row_key)

fig = go.Figure()
fig.add_trace(go.Scatter(
    x=cos, y=vals_prev,
    mode="lines+markers",
    name=f"{PREV_YEAR}년",
    line=dict(color="#5B9BD5", width=2),
    marker=dict(size=8),
    connectgaps=False,
))
fig.add_trace(go.Scatter(
    x=cos, y=vals_cur,
    mode="lines+markers",
    name=f"{RECENT_YEAR}년",
    line=dict(color="#ED7D31", width=2),
    marker=dict(size=8),
    connectgaps=False,
))
fig.update_layout(
    yaxis_title="인원 (명)",
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    hovermode="x unified",
    margin=dict(l=0, r=0, t=10, b=0),
    xaxis=dict(tickangle=-30),
    plot_bgcolor="white",
    yaxis=dict(gridcolor="#EEEEEE"),
)
st.plotly_chart(fig, use_container_width=True)

st.divider()

# ── Table 2 ───────────────────────────────────────────────────────────────────
st.subheader("주요 증권사 4개년 인당생산성 및 임직원수 / 순영업수익 추이")

# 구분선 역할을 하는 빈 행 삽입 (시각적 그룹핑)
y0, y1, y2, y3 = TREND_YEARS

T2_SECTIONS = {
    "인당생산성(백만원/인)":    "인당생산성 (백만원/인)",
    f"임직원수 {y0}년":         f"임직원수  {y0}년",
    f"임직원수 {y1}년":         f"임직원수  {y1}년",
    f"임직원수 {y2}년":         f"임직원수  {y2}년",
    f"임직원수 {y3}년":         f"임직원수  {y3}년",
    "임직원수 평균":             "임직원수  평균",
    f"순영업수익 {y0}년(억원)":  f"순영업수익  {y0}년 (억원)",
    f"순영업수익 {y1}년(억원)":  f"순영업수익  {y1}년 (억원)",
    f"순영업수익 {y2}년(억원)":  f"순영업수익  {y2}년 (억원)",
    f"순영업수익 {y3}년(억원)":  f"순영업수익  {y3}년 (억원)",
    "순영업수익 평균(억원)":      "순영업수익  평균 (억원)",
}
t2_fmt = format_df(add_group_averages(t2_raw)).rename(index=T2_SECTIONS)
t2_fmt = _mark_partial(t2_fmt, fin_data)
st.html(_render_table(t2_fmt, section_rows=True))

st.caption("**순영업수익 계산식**: 세전이익 − 영업외손익 + 판관비")
st.caption("**인당생산성**: 순영업수익 평균(백만원) ÷ 임직원수 평균 (조회기간 평균 기준)")
st.caption("**\\*** 판관비 미공시로 판관비 제외 계산 (세전이익 − 영업외손익)")

st.divider()

# ── 엑셀 다운로드 ──────────────────────────────────────────────────────────────
def _build_excel_bytes(fmt1: pd.DataFrame, fmt2: pd.DataFrame, fmt3: pd.DataFrame) -> bytes:
    import io
    import re as _re
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── 스타일 팔레트 ──────────────────────────────────────────────────────────
    def _fill(hex_): return PatternFill("solid", fgColor=hex_)
    F_YL   = _fill("FFF8DC")  # 우리투자증권
    F_GH   = _fill("D6EAF8")  # 그룹 헤더
    F_TH   = _fill("F2F3F4")  # 일반 헤더
    F_IDX  = _fill("F2F3F4")  # 인덱스 헤더
    F_AVG  = _fill("EEF2F7")  # 그룹 평균
    F_PROD = _fill("E9F7EF")  # 인당생산성
    F_SEC  = _fill("EBF5FB")  # 섹션 셀

    _thin = Side(style="thin", color="CCCCCC")
    _bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _c(ws, r, col, val="", fill=None, bold=False, italic=False, align="center"):
        cell = ws.cell(row=r, column=col, value=str(val) if val != "" else "")
        cell.border = _bdr
        if fill:
            cell.fill = fill
        cell.font = Font(bold=bold, italic=italic, size=10,
                         color="FFFFFF" if fill == F_GH and bold else "000000")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        return cell

    def _auto_width(ws):
        for col in ws.columns:
            w = max((len(str(cell.value or "")) for cell in col), default=6)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 22)

    # ── 시트 공통: 2행 헤더 + 데이터 ─────────────────────────────────────────
    def _write_t1(ws, df: pd.DataFrame):
        companies = list(df.columns)

        # 그룹 묶기
        grp_blocks: list[list] = []
        for co in companies:
            g = COMPANY_GROUP.get(co, "")
            if grp_blocks and grp_blocks[-1][0] == g:
                grp_blocks[-1][1].append(co)
            else:
                grp_blocks.append([g, [co]])

        # 헤더 행 1: 인덱스 + 그룹명
        _c(ws, 1, 1, "구분", fill=F_IDX, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        col = 2
        for g, cos in grp_blocks:
            if not g:
                for co in cos:
                    fill = F_YL if co == HIGHLIGHT_CO else F_TH
                    _c(ws, 1, col, co, fill=fill, bold=True)
                    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
                    col += 1
            else:
                _c(ws, 1, col, g, fill=F_GH, bold=True)
                if len(cos) > 1:
                    ws.merge_cells(start_row=1, start_column=col,
                                   end_row=1, end_column=col + len(cos) - 1)
                for co in cos:
                    fill = F_AVG if co in AVG_COLS else (F_YL if co == HIGHLIGHT_CO else F_TH)
                    disp = "평균" if co in AVG_COLS else co
                    _c(ws, 2, col, disp, fill=fill, bold=True, italic=(co in AVG_COLS))
                    col += 1

        # 데이터 행
        for ri, lbl in enumerate(df.index, start=3):
            _c(ws, ri, 1, lbl, bold=True, align="left")
            for ci, co in enumerate(companies, start=2):
                val = df.loc[lbl, co]
                fill = F_AVG if co in AVG_COLS else (F_YL if co == HIGHLIGHT_CO else None)
                _c(ws, ri, ci, val, fill=fill, italic=(co in AVG_COLS), align="right")

        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = ws.cell(row=3, column=2)
        _auto_width(ws)

    def _pre_fill(ws, r1, c1, r2, c2, fill):
        """merge 전 영역 전체에 fill/border 적용 (MergedCell 쓰기 오류 방지)."""
        for ri in range(r1, r2 + 1):
            for ci in range(c1, c2 + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.fill = fill
                cell.border = _bdr

    def _write_t2(ws, df: pd.DataFrame):
        companies = list(df.columns)

        grp_blocks: list[list] = []
        for co in companies:
            g = COMPANY_GROUP.get(co, "")
            if grp_blocks and grp_blocks[-1][0] == g:
                grp_blocks[-1][1].append(co)
            else:
                grp_blocks.append([g, [co]])

        # 헤더 행 1: 인덱스(2열) + 그룹명 — merge 전에 전체 영역 fill 적용
        _pre_fill(ws, 1, 1, 2, 2, F_IDX)
        _c(ws, 1, 1, "구분", fill=F_IDX, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)

        col = 3
        for g, cos in grp_blocks:
            if not g:
                for co in cos:
                    fill = F_YL if co == HIGHLIGHT_CO else F_TH
                    _c(ws, 1, col, co, fill=fill, bold=True)
                    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
                    col += 1
            else:
                _c(ws, 1, col, g, fill=F_GH, bold=True)
                if len(cos) > 1:
                    ws.merge_cells(start_row=1, start_column=col,
                                   end_row=1, end_column=col + len(cos) - 1)
                for co in cos:
                    fill = F_AVG if co in AVG_COLS else (F_YL if co == HIGHLIGHT_CO else F_TH)
                    disp = "평균" if co in AVG_COLS else co
                    _c(ws, 2, col, disp, fill=fill, bold=True, italic=(co in AVG_COLS))
                    col += 1

        def _sec(lbl):
            if "임직원수" in lbl:   return "임직원수"
            if "순영업수익" in lbl: return "순영업수익"
            return ""

        SEC_DISPLAY = {"임직원수": "임직원수 (명)", "순영업수익": "순영업수익 (억원)"}

        def _disp(lbl, sec):
            if sec in ("임직원수", "순영업수익"):
                if "평균" in lbl: return "평균"
                m = _re.search(r"(\d{4})", lbl)
                return f"{m.group(1)[2:]}년 12월" if m else lbl
            return lbl

        labels = list(df.index)
        sec_groups: list[list] = []
        for lbl in labels:
            s = _sec(lbl)
            if sec_groups and sec_groups[-1][0] == s:
                sec_groups[-1][1].append(lbl)
            else:
                sec_groups.append([s, [lbl]])

        r = 3
        for sec, lbls in sec_groups:
            is_prod = (sec == "")
            sec_start = r
            for j, lbl in enumerate(lbls):
                if is_prod:
                    # merge 전 두 셀 모두 fill 적용
                    _pre_fill(ws, r, 1, r, 2, F_PROD)
                    _c(ws, r, 1, _disp(lbl, sec), fill=F_PROD, bold=True, align="left")
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                else:
                    if j == 0:
                        # merge 전 섹션 전체 행의 열1 fill 적용
                        _pre_fill(ws, r, 1, r + len(lbls) - 1, 1, F_SEC)
                        _c(ws, r, 1, SEC_DISPLAY.get(sec, sec), fill=F_SEC, bold=True)
                        if len(lbls) > 1:
                            ws.merge_cells(start_row=r, start_column=1,
                                           end_row=r + len(lbls) - 1, end_column=1)
                    # j > 0 이면 열1은 MergedCell — 건드리지 않음
                    _c(ws, r, 2, _disp(lbl, sec), bold=True, align="left")

                for ci, co in enumerate(companies, start=3):
                    val = df.loc[lbl, co]
                    fill = (F_PROD if is_prod else
                            F_AVG  if co in AVG_COLS else
                            F_YL   if co == HIGHLIGHT_CO else None)
                    _c(ws, r, ci, val, fill=fill, italic=(co in AVG_COLS), align="right")
                r += 1

        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = ws.cell(row=3, column=3)
        _auto_width(ws)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"인력현황_{RECENT_YEAR}"
    _write_t1(ws1, fmt1)

    ws2 = wb.create_sheet(f"인력현황_{PREV_YEAR}")
    _write_t1(ws2, fmt2)

    ws3 = wb.create_sheet("추이분석")
    _write_t2(ws3, fmt3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_fmt1_xl = _merge_pct_rows(format_df(add_group_averages(t1_raw))).rename(index=T1_ROW_LABELS)
_fmt2_xl = _merge_pct_rows(format_df(add_group_averages(t1_raw_prev))).rename(index=T1_ROW_LABELS)

st.download_button(
    label="📥 엑셀 다운로드",
    data=_build_excel_bytes(_fmt1_xl, _fmt2_xl, t2_fmt),
    file_name=f"증권사_인력현황_{RECENT_YEAR}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=False,
)
