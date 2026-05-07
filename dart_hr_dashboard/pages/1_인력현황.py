import sys
import os
import re as _re

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from _shared import (
    HIGHLIGHT_CO,
    COMPANY_GROUP, AVG_COLS, AVG_COL_NAME, GROUPS_DEF,
    T1_ROW_LABELS,
    format_df, _merge_pct_rows, add_group_averages, _render_table,
    _apply_woori_override,
    _load, _load_t3, render_sidebar, render_debug,
)
from data_parser import (
    COMPANIES, PREV_YEAR, RECENT_YEAR, TREND_YEARS,
    build_table1, build_table2,
)

woori_ov = render_sidebar()

st.title("증권사 인력 현황 비교 대시보드")
st.caption(
    f"기준: {RECENT_YEAR}년 사업보고서 | 개별재무제표 | DART OpenAPI  ·  "
    f"노란색 열 = {HIGHLIGHT_CO}"
)

emp_data, fin_data, _src_main = _load()
t3_emp, t3_fin, t3_exec, _src_t3 = _load_t3()

render_debug(emp_data, t3_emp, _src_main, _src_t3)

t1_raw      = build_table1(emp_data, fin_data, RECENT_YEAR)
t1_raw_prev = build_table1(emp_data, fin_data, PREV_YEAR)
t2_raw      = build_table2(emp_data, fin_data)
t1_raw, t1_raw_prev, t2_raw = _apply_woori_override(t1_raw, t1_raw_prev, t2_raw, woori_ov)

def _build_t1_html(raw):
    fmt = _merge_pct_rows(format_df(add_group_averages(raw))).rename(index=T1_ROW_LABELS)
    return _render_table(fmt)

# ── Table 1 (최신연도) ────────────────────────────────────────────────────────
st.subheader(f"증권사별 자기자본 및 인력 현황 ({RECENT_YEAR}년)")
st.html(_build_t1_html(t1_raw))

st.divider()

# ── Table 2 (전년도) ──────────────────────────────────────────────────────────
st.subheader(f"증권사별 자기자본 및 인력 현황 ({PREV_YEAR}년)")
st.html(_build_t1_html(t1_raw_prev))

st.divider()

# ── 인력 비교 꺾은선 그래프 ──────────────────────────────────────────────────
st.subheader(f"{PREV_YEAR} vs {RECENT_YEAR} 인력 현황 비교")

METRIC_OPTIONS = {
    "총원":     "총인원",
    "리테일":   "리테일 인원",
    "본사영업": "본사영업 인원",
    "본사관리": "본사관리 인원",
}
selected_metric = st.radio(
    "기준", list(METRIC_OPTIONS.keys()),
    horizontal=True, label_visibility="collapsed",
)
row_key = METRIC_OPTIONS[selected_metric]
cos = list(COMPANIES.keys())

def _get_vals(raw: pd.DataFrame, key: str) -> list:
    if key not in raw.index:
        return [None] * len(cos)
    return [
        (float(raw.loc[key, co])
         if raw.loc[key, co] is not None
         and not (isinstance(raw.loc[key, co], float) and pd.isna(raw.loc[key, co]))
         else None)
        for co in cos
    ]

vals_prev = _get_vals(t1_raw_prev, row_key)
vals_cur  = _get_vals(t1_raw, row_key)

fig = go.Figure()
fig.add_trace(go.Scatter(
    x=cos, y=vals_prev, mode="lines+markers", name=f"{PREV_YEAR}년",
    line=dict(color="#5B9BD5", width=2), marker=dict(size=8), connectgaps=False,
))
fig.add_trace(go.Scatter(
    x=cos, y=vals_cur, mode="lines+markers", name=f"{RECENT_YEAR}년",
    line=dict(color="#ED7D31", width=2), marker=dict(size=8), connectgaps=False,
))
fig.update_layout(
    yaxis_title="인원 (명)",
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    hovermode="x unified",
    margin=dict(l=0, r=0, t=10, b=0),
    xaxis=dict(tickangle=-30),
    plot_bgcolor="white",
    yaxis=dict(gridcolor="#EEEEEE"),
)
st.plotly_chart(fig, width='stretch')

st.divider()

# ── 엑셀 다운로드 ──────────────────────────────────────────────────────────────
def _build_excel_bytes(fmt1: pd.DataFrame, fmt2: pd.DataFrame, fmt3: pd.DataFrame) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _fill(hex_): return PatternFill("solid", fgColor=hex_)
    F_YL      = _fill("FFF8DC")
    F_GH      = _fill("D6EAF8")
    F_TH      = _fill("F2F3F4")
    F_IDX     = _fill("F2F3F4")
    F_AVG     = _fill("EEF2F7")
    F_PROD    = _fill("E9F7EF")
    F_SEC     = _fill("EBF5FB")
    F_SEC_HDR = _fill("EDF4FC")
    F_SUB_REG = _fill("E8F8F5")
    F_SUB_CNT = _fill("F5EEF8")

    _thin = Side(style="thin", color="CCCCCC")
    _bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _c(ws, r, col, val="", fill=None, bold=False, italic=False, align="center"):
        cell = ws.cell(row=r, column=col, value=str(val) if val != "" else "")
        cell.border = _bdr
        if fill:
            cell.fill = fill
        cell.font = Font(bold=bold, italic=italic, size=10,
                         color="FFFFFF" if fill == F_GH and bold else "000000")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        return cell

    def _auto_width(ws):
        for col in ws.columns:
            w = max((len(str(cell.value or "")) for cell in col), default=6)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 22)

    def _write_t1(ws, df: pd.DataFrame):
        companies = list(df.columns)
        grp_blocks: list[list] = []
        for co in companies:
            g = COMPANY_GROUP.get(co, "")
            if grp_blocks and grp_blocks[-1][0] == g:
                grp_blocks[-1][1].append(co)
            else:
                grp_blocks.append([g, [co]])

        _c(ws, 1, 1, "구분", fill=F_IDX, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        col = 2
        for g, cos in grp_blocks:
            if not g:
                for co in cos:
                    fill = F_YL if co == HIGHLIGHT_CO else F_TH
                    _c(ws, 1, col, co, fill=fill, bold=True)
                    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
                    col += 1
            else:
                _c(ws, 1, col, g, fill=F_GH, bold=True)
                if len(cos) > 1:
                    ws.merge_cells(start_row=1, start_column=col,
                                   end_row=1, end_column=col + len(cos) - 1)
                for co in cos:
                    fill = F_AVG if co in AVG_COLS else (F_YL if co == HIGHLIGHT_CO else F_TH)
                    disp = "평균" if co in AVG_COLS else co
                    _c(ws, 2, col, disp, fill=fill, bold=True, italic=(co in AVG_COLS))
                    col += 1

        _XL_SEC_HDRS = {"리테일 인원", "본사영업 인원", "본사관리 인원"}

        def _xl_sub_kind(lbl):
            for sec in ("리테일", "본사영업", "본사관리"):
                for et in ("정규직", "기간제"):
                    if lbl == f"{sec} {et} 인원":
                        return et
            return None

        for ri, lbl in enumerate(df.index, start=3):
            kind = _xl_sub_kind(lbl)
            if kind:
                sub_fill = F_SUB_REG if kind == "정규직" else F_SUB_CNT
                _c(ws, ri, 1, f"  ▸ {kind}", fill=sub_fill, bold=False, align="left")
                for ci, co in enumerate(companies, start=2):
                    val  = df.loc[lbl, co]
                    fill = F_AVG if co in AVG_COLS else (F_YL if co == HIGHLIGHT_CO else sub_fill)
                    _c(ws, ri, ci, val, fill=fill, italic=(co in AVG_COLS), align="right")
                ws.row_dimensions[ri].height = 14
            elif lbl in _XL_SEC_HDRS:
                _c(ws, ri, 1, lbl, fill=F_SEC_HDR, bold=True, align="left")
                for ci, co in enumerate(companies, start=2):
                    val  = df.loc[lbl, co]
                    fill = F_AVG if co in AVG_COLS else (F_YL if co == HIGHLIGHT_CO else F_SEC_HDR)
                    _c(ws, ri, ci, val, fill=fill, italic=(co in AVG_COLS), align="right")
            else:
                _c(ws, ri, 1, lbl, bold=True, align="left")
                for ci, co in enumerate(companies, start=2):
                    val  = df.loc[lbl, co]
                    fill = F_AVG if co in AVG_COLS else (F_YL if co == HIGHLIGHT_CO else None)
                    _c(ws, ri, ci, val, fill=fill, italic=(co in AVG_COLS), align="right")

        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = ws.cell(row=3, column=2)
        _auto_width(ws)

    def _pre_fill(ws, r1, c1, r2, c2, fill):
        for ri in range(r1, r2 + 1):
            for ci in range(c1, c2 + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.fill = fill
                cell.border = _bdr

    def _write_t2(ws, df: pd.DataFrame):
        companies = list(df.columns)
        grp_blocks: list[list] = []
        for co in companies:
            g = COMPANY_GROUP.get(co, "")
            if grp_blocks and grp_blocks[-1][0] == g:
                grp_blocks[-1][1].append(co)
            else:
                grp_blocks.append([g, [co]])

        _pre_fill(ws, 1, 1, 2, 2, F_IDX)
        _c(ws, 1, 1, "구분", fill=F_IDX, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)

        col = 3
        for g, cos in grp_blocks:
            if not g:
                for co in cos:
                    fill = F_YL if co == HIGHLIGHT_CO else F_TH
                    _c(ws, 1, col, co, fill=fill, bold=True)
                    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
                    col += 1
            else:
                _c(ws, 1, col, g, fill=F_GH, bold=True)
                if len(cos) > 1:
                    ws.merge_cells(start_row=1, start_column=col,
                                   end_row=1, end_column=col + len(cos) - 1)
                for co in cos:
                    fill = F_AVG if co in AVG_COLS else (F_YL if co == HIGHLIGHT_CO else F_TH)
                    disp = "평균" if co in AVG_COLS else co
                    _c(ws, 2, col, disp, fill=fill, bold=True, italic=(co in AVG_COLS))
                    col += 1

        def _sec(lbl):
            if "임직원수" in lbl:   return "임직원수"
            if "순영업수익" in lbl: return "순영업수익"
            return ""

        SEC_DISPLAY = {"임직원수": "임직원수 (명)", "순영업수익": "순영업수익 (억원)"}

        def _disp(lbl, sec):
            if sec in ("임직원수", "순영업수익"):
                if "평균" in lbl: return "평균"
                m = _re.search(r"(\d{4})", lbl)
                return f"{m.group(1)[2:]}년 12월" if m else lbl
            return lbl

        labels = list(df.index)
        sec_groups: list[list] = []
        for lbl in labels:
            s = _sec(lbl)
            if sec_groups and sec_groups[-1][0] == s:
                sec_groups[-1][1].append(lbl)
            else:
                sec_groups.append([s, [lbl]])

        r = 3
        for sec, lbls in sec_groups:
            is_prod = (sec == "")
            for j, lbl in enumerate(lbls):
                if is_prod:
                    _pre_fill(ws, r, 1, r, 2, F_PROD)
                    _c(ws, r, 1, _disp(lbl, sec), fill=F_PROD, bold=True, align="left")
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                else:
                    if j == 0:
                        _pre_fill(ws, r, 1, r + len(lbls) - 1, 1, F_SEC)
                        _c(ws, r, 1, SEC_DISPLAY.get(sec, sec), fill=F_SEC, bold=True)
                        if len(lbls) > 1:
                            ws.merge_cells(start_row=r, start_column=1,
                                           end_row=r + len(lbls) - 1, end_column=1)
                    _c(ws, r, 2, _disp(lbl, sec), bold=True, align="left")

                for ci, co in enumerate(companies, start=3):
                    val = df.loc[lbl, co]
                    fill = (F_PROD if is_prod else
                            F_AVG  if co in AVG_COLS else
                            F_YL   if co == HIGHLIGHT_CO else None)
                    _c(ws, r, ci, val, fill=fill, italic=(co in AVG_COLS), align="right")
                r += 1

        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = ws.cell(row=3, column=3)
        _auto_width(ws)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"인력현황_{RECENT_YEAR}"
    _write_t1(ws1, fmt1)

    ws2 = wb.create_sheet(f"인력현황_{PREV_YEAR}")
    _write_t1(ws2, fmt2)

    ws3 = wb.create_sheet("추이분석")
    _write_t2(ws3, fmt3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_fmt1_xl = _merge_pct_rows(format_df(add_group_averages(t1_raw))).rename(index=T1_ROW_LABELS)
_fmt2_xl = _merge_pct_rows(format_df(add_group_averages(t1_raw_prev))).rename(index=T1_ROW_LABELS)

y0, y1, y2, y3 = TREND_YEARS
T2_SECTIONS = {
    "인당생산성(백만원/인)":    "인당생산성 (백만원/인)",
    f"임직원수 {y0}년":         f"임직원수  {y0}년",
    f"임직원수 {y1}년":         f"임직원수  {y1}년",
    f"임직원수 {y2}년":         f"임직원수  {y2}년",
    f"임직원수 {y3}년":         f"임직원수  {y3}년",
    "임직원수 평균":             "임직원수  평균",
    f"순영업수익 {y0}년(억원)":  f"순영업수익  {y0}년 (억원)",
    f"순영업수익 {y1}년(억원)":  f"순영업수익  {y1}년 (억원)",
    f"순영업수익 {y2}년(억원)":  f"순영업수익  {y2}년 (억원)",
    f"순영업수익 {y3}년(억원)":  f"순영업수익  {y3}년 (억원)",
    "순영업수익 평균(억원)":      "순영업수익  평균 (억원)",
}
_t2_fmt_xl = format_df(add_group_averages(t2_raw)).rename(index=T2_SECTIONS)

st.download_button(
    label="📥 엑셀 다운로드",
    data=_build_excel_bytes(_fmt1_xl, _fmt2_xl, _t2_fmt_xl),
    file_name=f"증권사_인력현황_{RECENT_YEAR}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    width='content',
)
